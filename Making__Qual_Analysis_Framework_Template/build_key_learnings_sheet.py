from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FILE_PATH = r"C:\Users\hyang\Documents\FuturesSummit_2025_QualStudy\Making__Qual_Analysis_Framework_Template\Qual_Analysis_Framework.xlsx"
SHEET_NAME = "Key_Learnings"
MAX_ITEMS = 15
DATA_START_ROW = 12
DATA_END_ROW = DATA_START_ROW + MAX_ITEMS - 1  # 26 when MAX_ITEMS=15

BUCKET_NAME = "Key Learnings"

BUCKET_DEFINITION = (
    "Key Learnings are substantive takeaways that emerge from patterns across multiple conference sessions "
    "or are articulated particularly clearly and concretely within a session. A Key Learning reflects something "
    "participants collectively learned about AI practice, governance, pedagogy, workforce preparation, or "
    "implementation—not a slogan or opinion, but a meaningful insight supported by evidence in the transcripts.\n\n"
    "Key Learnings are grounded entirely in what speakers and participants stated during sessions and are identified "
    "by synthesizing recurring ideas, shared conclusions, or clearly articulated lessons across sessions."
)

NOTEBOOKLM_PROMPT = (
    "Using only the provided session transcripts, identify key learnings that emerge across the conference sessions.\n\n"
    "Define a “key learning” as a substantive insight or lesson that:\n"
    "- Appears across multiple sessions, OR\n"
    "- Is articulated with particular clarity, specificity, or practical relevance in one or more sessions.\n\n"
    "For each key learning identified, provide:\n"
    "- A concise learning statement\n"
    "- Session_IDs where it appears\n"
    "- Speaker role(s) associated with the learning\n"
    "- 1–2 supporting quotes with Session_ID references\n\n"
    "Do not summarize individual sessions. Do not introduce external interpretation or opinion. "
    "Base all key learnings strictly on transcript evidence."
)

# Table columns (you asked to split evidence into multiple columns)
HEADERS = [

    "Key_Learning_ID",
    "Key_Learning_Statement",
    "Description_Notes",
    "Sessions_Appears_In (IDs)",
    "Confidence (High/Med/Low)",

    # Evidence split into multiple columns
    "Evidence_1_Session_ID",
    "Evidence_1_Speaker_Name",
    "Evidence_1_Speaker_Role",
    "Evidence_1_Quote_or_Excerpt",
    "Evidence_1_Why_This_Evidence_Matters",

    "Evidence_2_Session_ID",
    "Evidence_2_Speaker_Name",
    "Evidence_2_Speaker_Role",
    "Evidence_2_Quote_or_Excerpt",
    "Evidence_2_Why_This_Evidence_Matters",

    "Analyst_Notes",
    "Last_Updated"
]

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def main():
    wb = load_workbook(FILE_PATH)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f'Sheet "{SHEET_NAME}" not found in workbook. Current sheets: {wb.sheetnames}')

    ws = wb[SHEET_NAME]

    # Clear existing content safely (optional: keeps sheet but clears cells)
    ws.delete_rows(1, ws.max_row)

    # Styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(vertical="center")
    header_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
    section_fill = PatternFill("solid", fgColor="D9E1F2") # light blue

    # Row 1: Sheet title
    ws["A1"] = f"{BUCKET_NAME} — Analysis Framework"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    # Row 2-5: Definition section
    ws["A2"] = "Bucket Definition"
    ws["A2"].font = Font(bold=True)
    ws["A2"].fill = section_fill
    ws["A2"].alignment = center
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))

    ws["A3"] = BUCKET_DEFINITION
    ws["A3"].alignment = wrap
    ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=len(HEADERS))

    # Row 6-9: Prompt section
    ws["A6"] = "NotebookLM Prompt"
    ws["A6"].font = Font(bold=True)
    ws["A6"].fill = section_fill
    ws["A6"].alignment = center
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=len(HEADERS))

    ws["A7"] = NOTEBOOKLM_PROMPT
    ws["A7"].alignment = wrap
    ws.merge_cells(start_row=7, start_column=1, end_row=9, end_column=len(HEADERS))

    # Row 11: Table headers
    header_row = 11
    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Freeze panes below header row
    ws.freeze_panes = ws["A12"]

    # Set column widths (tuned for readability)
    widths = [
        16, 16, 42, 28, 22, 18,
        16, 20, 18, 52, 28,
        16, 20, 18, 52, 28,
        24, 16
    ]
    set_col_widths(ws, widths)

    # Add data validation for Confidence column (High/Med/Low)
    # Find the Confidence column index
    conf_col = HEADERS.index("Confidence (High/Med/Low)") + 1
    conf_letter = get_column_letter(conf_col)
    dv = DataValidation(type="list", formula1='"High,Med,Low"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{conf_letter}{DATA_START_ROW}:{conf_letter}{DATA_END_ROW}")

    # Set wrap + vertical top for data entry area
    for r in range(DATA_START_ROW, DATA_END_ROW + 1):
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=r, column=c).alignment = wrap

    wb.save(FILE_PATH)
    print(f"Updated sheet '{SHEET_NAME}' in: {FILE_PATH}")

if __name__ == "__main__":
    main()
