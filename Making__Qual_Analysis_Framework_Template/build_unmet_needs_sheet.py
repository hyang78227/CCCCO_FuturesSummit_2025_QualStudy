from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================
# CONFIG
# ============================================================
FILE_PATH = r"C:\Users\hyang\Documents\FuturesSummit_2025_QualStudy\Making__Qual_Analysis_Framework_Template\Qual_Analysis_Framework.xlsx"
SHEET_NAME = "Unmet_Needs"

MAX_ITEMS = 15
DATA_START_ROW = 12
DATA_END_ROW = DATA_START_ROW + MAX_ITEMS - 1  # 26


# ============================================================
# BUCKET CONTENT (Bucket 4: Unmet Needs)
# ============================================================
BUCKET_NAME = "Unmet Needs"

BUCKET_DEFINITION = (
    "Unmet Needs are identified when participants repeatedly express needs, requests, or concerns that are not "
    "sufficiently addressed during sessions. These needs may be stated directly through questions or inferred when "
    "responses remain high-level, vague, or incomplete despite repeated inquiry.\n\n"
    "Unmet Needs are grounded in observed patterns within transcripts and are not based on external expectations or "
    "evaluator judgment."
)

NOTEBOOKLM_PROMPT = (
    "Using only the session transcripts, identify unmet needs expressed or implied by participants.\n\n"
    "Define an unmet need as a need that:\n"
    "- Is explicitly requested by audience members but not fully addressed, OR\n"
    "- Is implied by repeated questions, deflected answers, or consistently high-level discussion.\n\n"
    "For each unmet need identified, provide:\n"
    "- Description of the unmet need\n"
    "- Who appears to have the need (e.g., faculty, administrators, students, IT)\n"
    "- Session_IDs where the need appears\n"
    "- Evidence from transcripts (quotes or paraphrases)\n"
    "- Why this unmet need matters for future programming\n\n"
    "Base all conclusions strictly on transcript patterns. Do not introduce external standards or opinions."
)


# ============================================================
# TABLE HEADERS (NO Bucket_Name column; FIRST column = item identified)
# ============================================================
HEADERS = [
    "Unmet_Need_ID",          # MUST be first column
    "Unmet_Need",
    "Unmet_Need_Description (Optional detail/clarifier)",
    "Who_Has_This_Need (Stakeholder group)",
    "Sessions_Appears_In (IDs)",
    "Confidence (High/Med/Low)",

    # Evidence 1 (split columns)
    "Evidence_1_Session_ID",
    "Evidence_1_Speaker_Name (if known)",
    "Evidence_1_Speaker_Role",
    "Evidence_1_Quote_or_Excerpt",
    "Evidence_1_Why_It_Supports_This_Unmet_Need",

    # Evidence 2 (split columns)
    "Evidence_2_Session_ID",
    "Evidence_2_Speaker_Name (if known)",
    "Evidence_2_Speaker_Role",
    "Evidence_2_Quote_or_Excerpt",
    "Evidence_2_Why_It_Supports_This_Unmet_Need",

    "Analyst_Notes",
    "Last_Updated"
]


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    wb = load_workbook(FILE_PATH)

    # Create sheet if missing
    if SHEET_NAME not in wb.sheetnames:
        wb.create_sheet(SHEET_NAME)

    ws = wb[SHEET_NAME]

    # Clear existing content
    ws.delete_rows(1, ws.max_row)

    # Styles (matching your existing style)
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(vertical="center")
    header_fill = PatternFill("solid", fgColor="1F4E79")   # dark blue
    section_fill = PatternFill("solid", fgColor="D9E1F2")  # light blue

    total_cols = len(HEADERS)

    # Row 1: Title
    ws["A1"] = f"{BUCKET_NAME} — Analysis Framework"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # Rows 2–5: Definition block
    ws["A2"] = "Bucket Definition"
    ws["A2"].font = Font(bold=True)
    ws["A2"].fill = section_fill
    ws["A2"].alignment = center
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)

    ws["A3"] = BUCKET_DEFINITION
    ws["A3"].alignment = wrap
    ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=total_cols)

    # Rows 6–9: Prompt block
    ws["A6"] = "NotebookLM Prompt"
    ws["A6"].font = Font(bold=True)
    ws["A6"].fill = section_fill
    ws["A6"].alignment = center
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=total_cols)

    ws["A7"] = NOTEBOOKLM_PROMPT
    ws["A7"].alignment = wrap
    ws.merge_cells(start_row=7, start_column=1, end_row=9, end_column=total_cols)

    # Row 11: Table headers
    header_row = 11
    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Freeze pane at first data row
    ws.freeze_panes = ws["A12"]

    # Column widths (tuned for unmet needs)
    widths = [
        34, 14, 40, 26, 22, 18,
        16, 22, 18, 52, 30,
        16, 22, 18, 52, 30,
        24, 16
    ]
    set_col_widths(ws, widths)

    # Confidence dropdown ONLY for the 15-row entry area
    conf_col = HEADERS.index("Confidence (High/Med/Low)") + 1
    conf_letter = get_column_letter(conf_col)
    dv = DataValidation(type="list", formula1='"High,Med,Low"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{conf_letter}{DATA_START_ROW}:{conf_letter}{DATA_END_ROW}")

    # Wrap + top align ONLY for the 15-row entry area
    for r in range(DATA_START_ROW, DATA_END_ROW + 1):
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).alignment = wrap

    wb.save(FILE_PATH)
    print(f"Updated sheet '{SHEET_NAME}' in: {FILE_PATH} (rows {DATA_START_ROW}-{DATA_END_ROW} capped at {MAX_ITEMS})")


if __name__ == "__main__":
    main()
