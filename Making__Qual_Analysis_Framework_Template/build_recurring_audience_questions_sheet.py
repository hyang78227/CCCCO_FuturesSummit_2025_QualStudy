from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FILE_PATH = r"C:\Users\hyang\Documents\FuturesSummit_2025_QualStudy\Qual_Analysis_Framework.xlsx"
SHEET_NAME = "Recurring_Audience_Questions"
MAX_ITEMS = 15
DATA_START_ROW = 12
DATA_END_ROW = DATA_START_ROW + MAX_ITEMS - 1  # 26

BUCKET_NAME = "Recurring Audience Questions"

BUCKET_DEFINITION = (
    "Recurring Audience Questions represent patterns in what audience members repeatedly asked or raised across "
    "sessions. These questions signal areas of uncertainty, demand, concern, or unmet need and often reveal "
    "priorities more clearly than panelist responses alone.\n\n"
    "Questions are analyzed in clusters rather than as isolated instances to surface the underlying needs driving "
    "them. Clusters must be grounded strictly in transcript evidence."
)

NOTEBOOKLM_PROMPT = (
    "Extract and analyze audience questions across all session transcripts.\n\n"
    "Group questions into recurring clusters based on shared intent or underlying need.\n\n"
    "For each question cluster, provide:\n"
    "- Cluster label\n"
    "- Description of the underlying need or concern\n"
    "- Session_IDs where the question appears\n"
    "- 1–2 example question excerpts or paraphrases with Session_ID references\n\n"
    "Do not list every individual question. Focus on recurring patterns grounded in transcript evidence. "
    "Use only the provided transcripts."
)

HEADERS = [
    "Question_Cluster_ID",
    "Cluster_Label",
    "Underlying_Need_or_Concern",
    "What_Audience_Is_Trying_To_Achieve",
    "Sessions_Appears_In (IDs)",
    "Confidence (High/Med/Low)",

    # Evidence (split into multiple columns)
    "Evidence_1_Session_ID",
    "Evidence_1_Question_As_Worded_or_Excerpt",
    "Evidence_1_Who_Asked (if known)",
    "Evidence_1_Why_It_Supports_This_Cluster",

    "Evidence_2_Session_ID",
    "Evidence_2_Question_As_Worded_or_Excerpt",
    "Evidence_2_Who_Asked (if known)",
    "Evidence_2_Why_It_Supports_This_Cluster",

    "Analyst_Notes",
    "Last_Updated"
]

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def main():
    wb = load_workbook(FILE_PATH)

    if SHEET_NAME not in wb.sheetnames:
        wb.create_sheet(SHEET_NAME)

    ws = wb[SHEET_NAME]
    ws.delete_rows(1, ws.max_row)

    # Styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(vertical="center")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    section_fill = PatternFill("solid", fgColor="D9E1F2")

    # Title
    ws["A1"] = f"{BUCKET_NAME} — Analysis Framework"
    ws["A1"].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    # Definition
    ws["A2"] = "Bucket Definition"
    ws["A2"].font = Font(bold=True)
    ws["A2"].fill = section_fill
    ws["A2"].alignment = center
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))

    ws["A3"] = BUCKET_DEFINITION
    ws["A3"].alignment = wrap
    ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=len(HEADERS))

    # Prompt
    ws["A6"] = "NotebookLM Prompt"
    ws["A6"].font = Font(bold=True)
    ws["A6"].fill = section_fill
    ws["A6"].alignment = center
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=len(HEADERS))

    ws["A7"] = NOTEBOOKLM_PROMPT
    ws["A7"].alignment = wrap
    ws.merge_cells(start_row=7, start_column=1, end_row=9, end_column=len(HEADERS))

    # Header row
    header_row = 11
    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    ws.freeze_panes = ws["A12"]

    # Column widths (tuned for questions)
    widths = [
        20, 18, 28, 40, 36, 22, 18,
        16, 52, 22, 30,
        16, 52, 22, 30,
        24, 16
    ]
    set_col_widths(ws, widths)

    # Confidence dropdown
    conf_col = HEADERS.index("Confidence (High/Med/Low)") + 1
    conf_letter = get_column_letter(conf_col)
    dv = DataValidation(type="list", formula1='"High,Med,Low"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{conf_letter}{DATA_START_ROW}:{conf_letter}{DATA_END_ROW}")

    # Wrap + top align for entry area
    for r in range(DATA_START_ROW, DATA_END_ROW + 1):
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=r, column=c).alignment = wrap

    wb.save(FILE_PATH)
    print(f"Updated sheet '{SHEET_NAME}' in: {FILE_PATH}")

if __name__ == "__main__":
    main()
