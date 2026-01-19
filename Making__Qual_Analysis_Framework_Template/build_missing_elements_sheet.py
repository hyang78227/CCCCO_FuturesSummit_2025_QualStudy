from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================
# CONFIG
# ============================================================
FILE_PATH = r"C:\Users\hyang\Documents\FuturesSummit_2025_QualStudy\Qual_Analysis_Framework.xlsx"
SHEET_NAME = "Missing_Elements"

MAX_ITEMS = 15
DATA_START_ROW = 12
DATA_END_ROW = DATA_START_ROW + MAX_ITEMS - 1  # 26


# ============================================================
# BUCKET CONTENT (Bucket 5: Missing Elements / Programming Gaps)
# ============================================================
BUCKET_NAME = "Missing Elements (Programming Gaps)"

BUCKET_DEFINITION = (
    "In this study, “missing elements” are identified by comparing what audience members asked for, "
    "what session topics implied they would cover, and how similar topics were addressed across sessions. "
    "An element is considered “missing” when a topic, use case, or detail is repeatedly requested or implied, "
    "but is not explained with sufficient depth, specificity, or concrete examples to be practically useful.\n\n"
    "Missing elements are therefore not based on external standards or personal judgment, but are grounded in patterns "
    "observed across session transcripts, audience questions, and differences in depth between stronger and weaker examples "
    "presented during the conference."
)

NOTEBOOKLM_PROMPT = (
    "Using only the provided session transcripts, identify “missing elements” across the conference sessions.\n\n"
    "Define a missing element as a topic, use case, example, or implementation detail that meets one or more of the following conditions:\n"
    "1) Audience members explicitly asked for it, but panel responses did not provide concrete, actionable detail.\n"
    "2) A session implied coverage of a topic (e.g., governance, training, use cases, workforce, evaluation), but did not explain it with sufficient depth, specificity, or examples.\n"
    "3) The topic appeared across multiple sessions, but was discussed inconsistently or only at a high level.\n"
    "4) One or two sessions provided strong, concrete examples, while other sessions addressing the same topic did not.\n\n"
    "For each missing element identified, provide:\n"
    "- A short description of what is missing\n"
    "- Which condition(s) above it meets (1–4)\n"
    "- Evidence from the transcripts (quotes or paraphrased statements with Session_ID references)\n"
    "- Why this missing element matters for future programming\n\n"
    "Do not introduce external standards or personal opinions. Base all observations strictly on patterns within the transcripts."
)


# ============================================================
# TABLE HEADERS
# - NO Bucket_Name column
# - FIRST column = item identified
# - Evidence split into multiple columns
# ============================================================
HEADERS = [
    "Missing_Element_ID",     # MUST be first
    "Missing_Element",
    "Missing_Element_Description (Optional detail/clarifier)",
    "Condition_Met (1-4)",
    "Topic_Area (Governance/Training/Use Cases/Workforce/Evaluation/Other)",
    "Sessions_Implicated (IDs)",
    "Confidence (High/Med/Low)",

    # Evidence 1
    "Evidence_1_Session_ID",
    "Evidence_1_Who (Speaker/Role)",
    "Evidence_1_Quote_or_Excerpt",
    "Evidence_1_Why_It_Shows_This_Is_Missing",

    # Evidence 2
    "Evidence_2_Session_ID",
    "Evidence_2_Who (Speaker/Role)",
    "Evidence_2_Quote_or_Excerpt",
    "Evidence_2_Why_It_Shows_This_Is_Missing",

    "Why_This_Matters_For_Future_Programming",
    "Suggested_Future_Session_or_Resource",
    "Analyst_Notes",
    "Last_Updated"
]


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    wb = load_workbook(FILE_PATH)

    # Create sheet if missing
    if SHEET_NAME not in wb.sheetnames:
        wb.create_sheet(SHEET_NAME)

    ws = wb[SHEET_NAME]

    # Clear existing content
    ws.delete_rows(1, ws.max_row)

    # Styles (match your earlier scripts)
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(vertical="center")
    header_fill = PatternFill("solid", fgColor="1F4E79")   # dark blue
    section_fill = PatternFill("solid", fgColor="D9E1F2")  # light blue

    total_cols = len(HEADERS)

    # Row 1: Title
    ws["A1"] = f"{BUCKET_NAME} — Analysis Framework"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # Rows 2–5: Definition block
    ws["A2"] = "Bucket Definition"
    ws["A2"].font = Font(bold=True)
    ws["A2"].fill = section_fill
    ws["A2"].alignment = center
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)

    ws["A3"] = BUCKET_DEFINITION
    ws["A3"].alignment = wrap
    ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=total_cols)

    # Rows 6–9: Prompt block
    ws["A6"] = "NotebookLM Prompt"
    ws["A6"].font = Font(bold=True)
    ws["A6"].fill = section_fill
    ws["A6"].alignment = center
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=total_cols)

    ws["A7"] = NOTEBOOKLM_PROMPT
    ws["A7"].alignment = wrap
    ws.merge_cells(start_row=7, start_column=1, end_row=9, end_column=total_cols)

    # Row 11: Table headers
    header_row = 11
    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Freeze at first data row
    ws.freeze_panes = ws["A12"]

    # Column widths (tuned)
    widths = [
        34, 16, 40, 14, 26, 22, 18,
        16, 22, 55, 30,
        16, 22, 55, 30,
        34, 34, 22, 16
    ]
    set_col_widths(ws, widths)

    # Confidence dropdown ONLY for rows 12–26
    conf_col = HEADERS.index("Confidence (High/Med/Low)") + 1
    conf_letter = get_column_letter(conf_col)
    dv = DataValidation(type="list", formula1='"High,Med,Low"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{conf_letter}{DATA_START_ROW}:{conf_letter}{DATA_END_ROW}")

    # Wrap + top align ONLY for the 15-row entry area
    for r in range(DATA_START_ROW, DATA_END_ROW + 1):
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).alignment = wrap

    wb.save(FILE_PATH)
    print(f"Updated sheet '{SHEET_NAME}' in: {FILE_PATH} (rows {DATA_START_ROW}-{DATA_END_ROW} capped at {MAX_ITEMS})")


if __name__ == "__main__":
    main()
