from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FILE_PATH = r"C:\Users\hyang\Documents\FuturesSummit_2025_QualStudy\Making__Qual_Analysis_Framework_Template\Qual_Analysis_Framework.xlsx"
SHEET_NAME = "Cross_Cutting_Themes"

BUCKET_NAME = "Cross-Cutting Themes"
MAX_ITEMS = 15
DATA_START_ROW = 12
DATA_END_ROW = DATA_START_ROW + MAX_ITEMS - 1  # 26

BUCKET_DEFINITION = (
    "Cross-Cutting Themes are organizing patterns that appear across multiple sessions, topics, or stakeholder "
    "perspectives. Themes represent shared areas of focus, tension, concern, or opportunity that recur throughout "
    "the conference, even when discussed in different contexts.\n\n"
    "Themes help explain what participants collectively grappled with and provide a structured way to organize "
    "diverse session content into coherent analytic categories. Themes must be grounded in transcript evidence and "
    "should not be based on external standards or personal judgment."
)

NOTEBOOKLM_PROMPT = (
    "Across all provided session transcripts, identify cross-cutting themes that appear across multiple sessions.\n\n"
    "For each theme identified, provide:\n"
    "- Theme name\n"
    "- A clear definition describing what the theme encompasses\n"
    "- Session_IDs where the theme appears\n"
    "- How the theme manifests differently across contexts or roles (if applicable)\n"
    "- 1–2 representative quotes with Session_ID references\n\n"
    "Use only transcript content. Do not summarize individual sessions. Do not introduce external interpretation "
    "or opinion. Base all themes strictly on transcript evidence."
)

# Evidence split into multiple columns (as you requested)
HEADERS = [
    "Theme_ID",
    "Theme_Name",
    "Theme_Definition",
    "How_Theme_Shows_Up_Across_Contexts",
    "Sessions_Appears_In (IDs)",
    "Confidence (High/Med/Low)",

    "Evidence_1_Session_ID",
    "Evidence_1_Speaker_Name",
    "Evidence_1_Speaker_Role",
    "Evidence_1_Quote_or_Excerpt",
    "Evidence_1_Why_This_Evidence_Matters",

    "Evidence_2_Session_ID",
    "Evidence_2_Speaker_Name",
    "Evidence_2_Speaker_Role",
    "Evidence_2_Quote_or_Excerpt",
    "Evidence_2_Why_This_Evidence_Matters",

    "Analyst_Notes",
    "Last_Updated"
]

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def main():
    wb = load_workbook(FILE_PATH)

    # Create sheet if missing
    if SHEET_NAME not in wb.sheetnames:
        wb.create_sheet(SHEET_NAME)

    ws = wb[SHEET_NAME]

    # Clear existing content
    ws.delete_rows(1, ws.max_row)
 
    #Freeze panes to keep header visible
    ws.freeze_panes = ws["A12"]


    # Styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(vertical="center")
    header_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
    section_fill = PatternFill("solid", fgColor="D9E1F2") # light blue

    # Row 1: Title
    ws["A1"] = f"{BUCKET_NAME} — Analysis Framework"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    # Definition block
    ws["A2"] = "Bucket Definition"
    ws["A2"].font = Font(bold=True)
    ws["A2"].fill = section_fill
    ws["A2"].alignment = center
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))

    ws["A3"] = BUCKET_DEFINITION
    ws["A3"].alignment = wrap
    ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=len(HEADERS))

    # Prompt block
    ws["A6"] = "NotebookLM Prompt"
    ws["A6"].font = Font(bold=True)
    ws["A6"].fill = section_fill
    ws["A6"].alignment = center
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=len(HEADERS))

    ws["A7"] = NOTEBOOKLM_PROMPT
    ws["A7"].alignment = wrap
    ws.merge_cells(start_row=7, start_column=1, end_row=9, end_column=len(HEADERS))

    # Table header row
    header_row = 11
    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    ws.freeze_panes = ws["A12"]

    # Column widths tuned for themes
    widths = [
        18, 12, 26, 40, 36, 22, 18,
        16, 20, 18, 52, 28,
        16, 20, 18, 52, 28,
        24, 16
    ]
    set_col_widths(ws, widths)

    # Confidence dropdown
    conf_col = HEADERS.index("Confidence (High/Med/Low)") + 1
    conf_letter = get_column_letter(conf_col)
    dv = DataValidation(type="list", formula1='"High,Med,Low"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{conf_letter}{DATA_START_ROW}:{conf_letter}{DATA_END_ROW}")

    # Wrap + top align for entry area
    for r in range(DATA_START_ROW, DATA_END_ROW + 1):
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=r, column=c).alignment = wrap

    wb.save(FILE_PATH)
    print(f"Updated sheet '{SHEET_NAME}' in: {FILE_PATH}")

if __name__ == "__main__":
    main()
