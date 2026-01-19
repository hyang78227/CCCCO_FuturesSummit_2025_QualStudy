from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================
# CONFIG
# ============================================================
FILE_PATH = r"C:\Users\hyang\Documents\FuturesSummit_2025_QualStudy\Qual_Analysis_Framework.xlsx"
SHEET_NAME = "Use_Cases_Exemplars"

MAX_ITEMS = 15
DATA_START_ROW = 12
DATA_END_ROW = DATA_START_ROW + MAX_ITEMS - 1  # 26


# ============================================================
# BUCKET CONTENT (Bucket 6: Use Cases & Exemplars)
# ============================================================
BUCKET_NAME = "Use Cases & Exemplars"

BUCKET_DEFINITION = (
    "Use Cases are concrete examples of how AI was applied in practice, describing who used AI, in what context, "
    "for what purpose, and with what outcome. Exemplars are particularly strong or well-articulated use cases that "
    "demonstrate clarity, replicability, or impact.\n\n"
    "Use cases and exemplars are identified solely from transcript content and are not generalized beyond what speakers described."
)

NOTEBOOKLM_PROMPT = (
    "Identify use cases and exemplars described in the session transcripts.\n\n"
    "For each use case identified, provide:\n"
    "- Use case title\n"
    "- Who is using AI (role)\n"
    "- Context and purpose\n"
    "- Benefits or outcomes described\n"
    "- Risks or constraints mentioned (if any)\n"
    "- Session_ID and speaker role\n"
    "- Supporting quotes\n\n"
    "Then identify which use cases function as exemplars and explain why (clarity, specificity, replicability, or impact).\n\n"
    "Do not invent details or infer beyond transcript evidence."
)


# ============================================================
# TABLE HEADERS
# - NO Bucket_Name column
# - FIRST column = item identified
# - Evidence split into multiple columns
# ============================================================
HEADERS = [
    "Use_Case_ID",                 # MUST be first
    "Use_Case",
    "Is_Exemplar (Y/N)",
    "Why_Exemplar (if Y)",
    "Intended_User (Role)",
    "Context",
    "Purpose / Job-to-be-done",
    "Workflow_Steps (as described)",
    "Tools_or_Systems_Mentioned",
    "Benefits_or_Outcomes_Claimed",
    "Risks_or_Constraints_Mentioned",
    "Sessions_Appears_In (IDs)",
    "Confidence (High/Med/Low)",

    # Evidence 1
    "Evidence_1_Session_ID",
    "Evidence_1_Who (Speaker/Role)",
    "Evidence_1_Quote_or_Excerpt",
    "Evidence_1_Why_It_Supports_This_Use_Case",

    # Evidence 2
    "Evidence_2_Session_ID",
    "Evidence_2_Who (Speaker/Role)",
    "Evidence_2_Quote_or_Excerpt",
    "Evidence_2_Why_It_Supports_This_Use_Case",

    "Clip_Candidate (Describe moment; no timestamps yet)",
    "Suggested_Future_Session_Topic",
    "Analyst_Notes",
    "Last_Updated"
]


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    wb = load_workbook(FILE_PATH)

    # Create sheet if missing
    if SHEET_NAME not in wb.sheetnames:
        wb.create_sheet(SHEET_NAME)

    ws = wb[SHEET_NAME]

    # Clear existing content
    ws.delete_rows(1, ws.max_row)

    # Styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(vertical="center")
    header_fill = PatternFill("solid", fgColor="1F4E79")   # dark blue
    section_fill = PatternFill("solid", fgColor="D9E1F2")  # light blue

    total_cols = len(HEADERS)

    # Row 1: Title
    ws["A1"] = f"{BUCKET_NAME} — Analysis Framework"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # Rows 2–5: Definition block
    ws["A2"] = "Bucket Definition"
    ws["A2"].font = Font(bold=True)
    ws["A2"].fill = section_fill
    ws["A2"].alignment = center
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)

    ws["A3"] = BUCKET_DEFINITION
    ws["A3"].alignment = wrap
    ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=total_cols)

    # Rows 6–9: Prompt block
    ws["A6"] = "NotebookLM Prompt"
    ws["A6"].font = Font(bold=True)
    ws["A6"].fill = section_fill
    ws["A6"].alignment = center
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=total_cols)

    ws["A7"] = NOTEBOOKLM_PROMPT
    ws["A7"].alignment = wrap
    ws.merge_cells(start_row=7, start_column=1, end_row=9, end_column=total_cols)

    # Row 11: Table headers
    header_row = 11
    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Freeze at first data row
    ws.freeze_panes = ws["A12"]

    # Column widths (wide sheet; tuned to stay readable)
    widths = [
        34, 12, 14, 22, 20, 22, 22, 28, 22, 24, 24, 18, 18,
        16, 22, 55, 30,
        16, 22, 55, 30,
        34, 26, 22, 16
    ]
    set_col_widths(ws, widths)

    # Confidence dropdown ONLY for rows 12–26
    conf_col = HEADERS.index("Confidence (High/Med/Low)") + 1
    conf_letter = get_column_letter(conf_col)
    dv = DataValidation(type="list", formula1='"High,Med,Low"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{conf_letter}{DATA_START_ROW}:{conf_letter}{DATA_END_ROW}")

    # Wrap + top align ONLY for the 15-row entry area
    for r in range(DATA_START_ROW, DATA_END_ROW + 1):
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).alignment = wrap

    wb.save(FILE_PATH)
    print(f"Updated sheet '{SHEET_NAME}' in: {FILE_PATH} (rows {DATA_START_ROW}-{DATA_END_ROW} capped at {MAX_ITEMS})")


if __name__ == "__main__":
    main()
