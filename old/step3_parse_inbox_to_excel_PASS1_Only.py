"""Step 3: Parse NotebookLM INBOX response .md files and append results into Qual_Analysis_Framework.xlsx.

This version is robust to NotebookLM outputs saved as Markdown that may contain:
- escaped underscores (e.g., RUN\_ID)
- escaped numbered items (e.g., 1\.)
- HTML non-breaking spaces (&nbsp;)
- an "outer" BEGIN/END wrapper and an "inner" escaped BEGIN/END wrapper

It writes a NEW Excel file and NEW Session_Index.csv so originals are not overwritten.
"""

import argparse
import csv
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


# -----------------------------
# Text normalization + extraction
# -----------------------------
OUTER_BEGIN = "---BEGIN_NOTEBOOKLM_RESPONSE---"
OUTER_END = "---END_NOTEBOOKLM_RESPONSE---"
INNER_BEGIN_ESC = "---BEGIN\\_NOTEBOOKLM\\_RESPONSE---"
INNER_END_ESC = "---END\\_NOTEBOOKLM\\_RESPONSE---"


def normalize_md(text: str) -> str:
    """Normalize common Markdown/HTML artifacts produced by copy/paste."""
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    # HTML entities
    text = text.replace("&nbsp;", " ")

    # Common backslash escapes from Markdown
    text = text.replace("\\_", "_")
    text = text.replace("\\.", ".")

    # Some copy/pastes include NBSP as a literal character
    text = text.replace("\u00a0", " ")

    # Normalize inner markers (escaped) to plain
    text = text.replace(INNER_BEGIN_ESC, OUTER_BEGIN)
    text = text.replace(INNER_END_ESC, OUTER_END)

    return text


def extract_response_block(full_text: str) -> str:
    """Extract the NotebookLM response content from a file.

    Supports:
    - outer BEGIN/END markers
    - or only inner escaped markers (normalized by normalize_md)
    - otherwise returns full text

    Also strips any header metadata that appears between a BEGIN marker and the start of numbered items.
    """
    t = normalize_md(full_text)

    if OUTER_BEGIN in t and OUTER_END in t:
        block = t.split(OUTER_BEGIN, 1)[1].split(OUTER_END, 1)[0]
    else:
        # Fallback: treat whole file as content
        block = t

    block = block.strip()

    # If the pasted content included header metadata lines (RUN_ID, PASS_ID, etc.)
    # remove them by trimming everything before the first numbered item.
    m = re.search(r"(?m)^\s*(?:\*\*\s*)?(\d+)\.\s+", block)
    if m:
        block = block[m.start():].strip()

    return block


def split_numbered_items(block: str) -> List[Tuple[int, str]]:
    """Split content into numbered items. Robust to optional bolding like **1. Title**."""
    block = normalize_md(block).strip()

    # Match starts like:
    # 1. ...
    # **1. ...**
    pattern = re.compile(r"(?m)^\s*(?:\*\*\s*)?(\d+)\.\s+")
    matches = list(pattern.finditer(block))
    if not matches:
        raise ValueError("No numbered items found (expected '1.', '2.', etc.).")

    items: List[Tuple[int, str]] = []
    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(block)
        num = int(m.group(1))
        item_text = block[m.end():end].strip()
        items.append((num, item_text))

    return items


def _strip_md_emphasis(s: str) -> str:
    # Remove bold/italic markers while keeping text
    s = s.replace("**", "")
    s = s.replace("*", "")
    return s


def find_field(item_text: str, label: str) -> str:
    """Extract a labeled field like 'Definition:' or '**Definition:**' from an item."""
    t = normalize_md(item_text)

    # Remove leading bullet symbols for easier parsing
    lines = []
    for line in t.split("\n"):
        line = line.strip()
        line = re.sub(r"^[\u2022\-\*\u25e6]+\s+", "", line)  # • - * ◦
        lines.append(line)
    t = "\n".join(lines)

    # Allow label variants like "Session_IDs" vs "Session IDs"
    label_pattern = re.escape(label).replace("_", "[_ ]")

    pattern = re.compile(
        rf"(?is)(?:^|\n)\s*(?:\*\*\s*)?{label_pattern}(?:\s*\*\*)?\s*:\s*(.*?)(?=\n\s*(?:\*\*\s*)?[A-Za-z0-9][A-Za-z0-9_ /()\-]*\s*(?:\*\*)?\s*:\s|\Z)"
    )

    m = pattern.search(t)
    if not m:
        return ""

    return _strip_md_emphasis(m.group(1)).strip()


def find_sessions(item_text: str) -> List[str]:
    s = find_field(item_text, "Session_IDs")
    if not s:
        return []
    return [p.strip() for p in s.split(",") if p.strip()]


def find_quote_lines(item_text: str) -> List[Dict[str, str]]:
    """Extract quote lines like: "..." (FS01) or "..."."""
    t = normalize_md(item_text)
    quotes: List[Dict[str, str]] = []
    for raw_line in t.split("\n"):
        line = raw_line.strip()
        if not line:
            continue
        line = re.sub(r"^[\u2022\-\*\u25e6]+\s+", "", line)
        m = re.search(r'"(.*?)"\s*(?:\((FS\d+)\))?\s*$', line)
        if m:
            quotes.append({"quote": m.group(1).strip(), "session_id": (m.group(2) or "").strip()})
    return quotes


def today_str() -> str:
    return datetime.now().strftime("%Y-%m-%d")


# -----------------------------
# Excel utilities
# -----------------------------
def get_headers(ws, header_row: int = 11) -> List[str]:
    headers = []
    col = 1
    while True:
        v = ws.cell(header_row, col).value
        if v is None:
            break
        headers.append(str(v).strip())
        col += 1
    return headers


def find_next_empty_row(ws, header_row: int = 11) -> int:
    row = header_row + 1
    while ws.cell(row, 1).value not in (None, ""):
        row += 1
    return row


def append_row(ws, headers: List[str], row_dict: Dict[str, str], header_row: int = 11) -> int:
    row_idx = find_next_empty_row(ws, header_row=header_row)
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row_idx, col_idx).value = row_dict.get(h, "")
    return row_idx


# -----------------------------
# Bucket-specific row builders
# -----------------------------
def build_cross_cutting_theme_row(run_id: str, pass_id: str, num: int, item_text: str) -> Dict[str, str]:
    theme_name = _strip_md_emphasis(item_text.split("\n", 1)[0].strip())
    theme_def = find_field(item_text, "Definition")
    contexts = find_field(item_text, "Contextual Differences")
    sessions = find_sessions(item_text)
    quotes = find_quote_lines(item_text)

    conf = "High" if len(sessions) >= 2 else "Medium"
    q1 = quotes[0] if len(quotes) > 0 else {"quote": "", "session_id": ""}
    q2 = quotes[1] if len(quotes) > 1 else {"quote": "", "session_id": ""}

    return {
        "Theme_ID": f"CCT_{num:02d}",
        "Theme_Name": theme_name,
        "Theme_Definition": theme_def,
        "How_Theme_Shows_Up_Across_Contexts": contexts,
        "Sessions_Appears_In (IDs)": ", ".join(sessions),
        "Confidence (High/Med/Low)": conf,
        "Evidence_1_Session_ID": q1["session_id"] or (sessions[0] if sessions else ""),
        "Evidence_1_Speaker_Name": "",
        "Evidence_1_Speaker_Role": "",
        "Evidence_1_Quote_or_Excerpt": q1["quote"],
        "Evidence_1_Why_This_Evidence_Matters": "",
        "Evidence_2_Session_ID": q2["session_id"] or (sessions[1] if len(sessions) > 1 else (sessions[0] if sessions else "")),
        "Evidence_2_Speaker_Name": "",
        "Evidence_2_Speaker_Role": "",
        "Evidence_2_Quote_or_Excerpt": q2["quote"],
        "Evidence_2_Why_This_Evidence_Matters": "",
        "Analyst_Notes": f"Imported from INBOX ({run_id}, {pass_id})",
        "Last_Updated": today_str(),
    }


def build_key_learning_row(run_id: str, pass_id: str, num: int, item_text: str) -> Dict[str, str]:
    title = _strip_md_emphasis(item_text.split("\n", 1)[0].strip())
    concise = find_field(item_text, "Concise learning statement")
    sessions = find_sessions(item_text)
    roles = find_field(item_text, "Speaker role(s)")
    quotes = find_quote_lines(item_text)

    conf = "High" if len(sessions) >= 2 else "Medium"
    q1 = quotes[0] if len(quotes) > 0 else {"quote": "", "session_id": ""}
    q2 = quotes[1] if len(quotes) > 1 else {"quote": "", "session_id": ""}

    ev1_sid = q1["session_id"] or (sessions[0] if sessions else "")
    ev2_sid = q2["session_id"] or (sessions[1] if len(sessions) > 1 else (sessions[0] if sessions else ""))

    return {
        "Key_Learning_ID": f"KL_{num:02d}",
        "Key_Learning_Statement": concise or title,
        "Description_Notes": title if concise else "",
        "Sessions_Appears_In (IDs)": ", ".join(sessions),
        "Confidence (High/Med/Low)": conf,
        "Evidence_1_Session_ID": ev1_sid,
        "Evidence_1_Speaker_Name": "",
        "Evidence_1_Speaker_Role": roles,
        "Evidence_1_Quote_or_Excerpt": q1["quote"],
        "Evidence_1_Why_This_Evidence_Matters": "",
        "Evidence_2_Session_ID": ev2_sid,
        "Evidence_2_Speaker_Name": "",
        "Evidence_2_Speaker_Role": roles,
        "Evidence_2_Quote_or_Excerpt": q2["quote"],
        "Evidence_2_Why_This_Evidence_Matters": "",
        "Analyst_Notes": f"Imported from INBOX ({run_id}, {pass_id})",
        "Last_Updated": today_str(),
    }


def build_recurring_questions_row(run_id: str, pass_id: str, num: int, item_text: str) -> Dict[str, str]:
    cluster_title = _strip_md_emphasis(item_text.split("\n", 1)[0].strip())
    cluster_label = find_field(item_text, "Cluster Label")
    underlying = find_field(item_text, "Description of the underlying need or concern")
    sessions = find_sessions(item_text)

    ex_block = find_field(item_text, "Example question excerpts")
    excerpts: List[str] = []
    if ex_block:
        for line in ex_block.split("\n"):
            m = re.search(r'"(.*?)"', line)
            if m:
                excerpts.append(m.group(1).strip())

    conf = "High" if len(sessions) >= 2 else "Medium"
    q1 = excerpts[0] if len(excerpts) > 0 else ""
    q2 = excerpts[1] if len(excerpts) > 1 else ""

    ev1_sid = sessions[0] if sessions else ""
    ev2_sid = sessions[1] if len(sessions) > 1 else (sessions[0] if sessions else "")

    return {
        "Question_Cluster_ID": f"QAQ_{num:02d}",
        "Cluster_Label": cluster_label or cluster_title,
        "Underlying_Need_or_Concern": underlying,
        "What_Audience_Is_Trying_To_Achieve": "",
        "Sessions_Appears_In (IDs)": ", ".join(sessions),
        "Confidence (High/Med/Low)": conf,
        "Evidence_1_Session_ID": ev1_sid,
        "Evidence_1_Question_As_Worded_or_Excerpt": q1,
        "Evidence_1_Who_Asked (if known)": "",
        "Evidence_1_Why_It_Supports_This_Cluster": "",
        "Evidence_2_Session_ID": ev2_sid,
        "Evidence_2_Question_As_Worded_or_Excerpt": q2,
        "Evidence_2_Who_Asked (if known)": "",
        "Evidence_2_Why_It_Supports_This_Cluster": "",
        "Analyst_Notes": f"Imported from INBOX ({run_id}, {pass_id})",
        "Last_Updated": today_str(),
    }


# -----------------------------
# Main
# -----------------------------
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--run_id", required=True)
    ap.add_argument("--pass_id", default="PASS_1")
    ap.add_argument("--project_root", default=".")
    ap.add_argument("--inbox_dir", default=None, help="Optional: override INBOX directory")
    args = ap.parse_args()

    project_root = Path(args.project_root).resolve()
    run_id = args.run_id
    pass_id = args.pass_id

    excel_path = project_root / "Qual_Analysis_Framework.xlsx"
    session_index_path = project_root / "Session_Index.csv"

    # INBOX location defaults to standard folder; if not present, fall back to project root
    if args.inbox_dir:
        inbox_dir = Path(args.inbox_dir).resolve()
    else:
        default_inbox = project_root / "03_NotebookLM_Exports" / "INBOX"
        inbox_dir = default_inbox if default_inbox.exists() else project_root

    if not excel_path.exists():
        raise FileNotFoundError(f"Missing: {excel_path}")
    if not session_index_path.exists():
        raise FileNotFoundError(f"Missing: {session_index_path}")

    needed = [
        ("Cross_Cutting_Themes", build_cross_cutting_theme_row),
        ("Key_Learnings", build_key_learning_row),
        ("Recurring_Audience_Questions", build_recurring_questions_row),
    ]

    wb = load_workbook(excel_path)
    processed_buckets: List[str] = []

    for bucket_name, row_builder in needed:
        inbox_file = inbox_dir / f"{run_id}__{pass_id}__{bucket_name}__NotebookLM_Response.md"
        if not inbox_file.exists():
            raise FileNotFoundError(f"Missing INBOX file: {inbox_file}")

        raw = inbox_file.read_text(encoding="utf-8")
        response_block = extract_response_block(raw)
        items = split_numbered_items(response_block)

        if bucket_name not in wb.sheetnames:
            raise KeyError(f"Excel is missing expected sheet: {bucket_name}")

        ws = wb[bucket_name]
        headers = get_headers(ws, header_row=11)

        for num, item_text in items:
            row_dict = row_builder(run_id, pass_id, num, item_text)
            append_row(ws, headers, row_dict, header_row=11)

        processed_buckets.append(bucket_name)
        print(f"[OK] {bucket_name}: appended {len(items)} row(s) from {inbox_file.name}")

    out_excel = project_root / f"Qual_Analysis_Framework__UPDATED_{run_id}.xlsx"
    wb.save(out_excel)
    print(f"[OK] Saved updated Excel: {out_excel.name}")

    out_csv = project_root / f"Session_Index__UPDATED_{run_id}.csv"

    with session_index_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = reader.fieldnames or []

    if "NotebookLM_Status" not in fieldnames:
        fieldnames.append("NotebookLM_Status")

    status_value = f"Pass1 registered to Excel ({run_id})"
    for r in rows:
        r["NotebookLM_Status"] = status_value

    with out_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"[OK] Saved updated Session_Index: {out_csv.name}")
    print(f"[DONE] Buckets processed: {', '.join(processed_buckets)}")


if __name__ == "__main__":
    main()
