import os
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook

# 1. Define the folder the script is in
SCRIPT_DIR = Path(__file__).resolve().parent 

# 2. Define the True Project Root (one level up)
PROJECT_ROOT = SCRIPT_DIR.parent 

# 3. Use PROJECT_ROOT for all main folders
EXCEL_PATH = PROJECT_ROOT / "Qual_Analysis_Framework.xlsx"
OUT_PROMPTS_DIR = PROJECT_ROOT / "03_NotebookLM_Exports" / "PROMPTS_READY_TO_PASTE"
OUT_INBOX_DIR = PROJECT_ROOT / "03_NotebookLM_Exports" / "INBOX"


NOTEBOOK_NAME = "FuturesSummit_2025_Qual_Analysis_AllSessions"
ANALYST = "Hongling"

# Choose your pass naming
PASS_ID = "PASS_1"

# Add this line to enforce machine-splittable output
NUMBERED_LIST_SUFFIX = "\n\nReturn the output as a numbered list (1., 2., 3., ...). One finding per number."

def find_prompt_cell(ws):
    """Find the cell containing 'NotebookLM Prompt' and return the cell below it (prompt text)."""
    for row in range(1, 50):
        for col in range(1, 10):
            val = ws.cell(row, col).value
            if isinstance(val, str) and val.strip() == "NotebookLM Prompt":
                return ws.cell(row + 1, col)  # prompt text cell
    return None

def main():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel not found: {EXCEL_PATH}")

    OUT_PROMPTS_DIR.mkdir(parents=True, exist_ok=True)
    OUT_INBOX_DIR.mkdir(parents=True, exist_ok=True)

    run_id = "RUN_" + datetime.now().strftime("%Y%m%d_%H%M")

    wb = load_workbook(EXCEL_PATH, data_only=False)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        prompt_cell = find_prompt_cell(ws)
        if prompt_cell is None:
            print(f"[WARN] No 'NotebookLM Prompt' found in sheet: {sheet_name}")
            continue

        prompt_text = (prompt_cell.value or "").strip()
        if not prompt_text:
            print(f"[WARN] Empty prompt text in sheet: {sheet_name}")
            continue

        # Create prompt file (ready to paste into NotebookLM)
        prompt_file = OUT_PROMPTS_DIR / f"{run_id}__{PASS_ID}__{sheet_name}__PROMPT.txt"
        prompt_file.write_text(prompt_text + NUMBERED_LIST_SUFFIX, encoding="utf-8")

        # Create INBOX response template file (where you paste NotebookLM answer)
        inbox_file = OUT_INBOX_DIR / f"{run_id}__{PASS_ID}__{sheet_name}__NotebookLM_Response.md"
        header = (
            f"RUN_ID: {run_id}\n"
            f"PASS_ID: {PASS_ID}\n"
            f"BUCKET: {sheet_name}\n"
            f"NOTEBOOK: {NOTEBOOK_NAME}\n"
            f"DATE: {datetime.now().strftime('%Y-%m-%d')}\n"
            f"ANALYST: {ANALYST}\n"
            f"SOURCE_SET: All transcripts in 02_Transcripts_Clean\n"
            f"---BEGIN_NOTEBOOKLM_RESPONSE---\n\n"
            f"(PASTE NOTEBOOKLM RESPONSE HERE)\n\n"
            f"---END_NOTEBOOKLM_RESPONSE---\n"
        )
        inbox_file.write_text(header, encoding="utf-8")

        print(f"[OK] Generated prompt + INBOX template for bucket: {sheet_name}")

    print("\nDONE.")
    print(f"Run ID: {run_id}")
    print(f"Prompt files: {OUT_PROMPTS_DIR}")
    print(f"INBOX templates: {OUT_INBOX_DIR}")

if __name__ == "__main__":
    main()
