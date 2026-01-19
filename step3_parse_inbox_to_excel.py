"""Step 3: Parse NotebookLM INBOX response .md files and append results into Qual_Analysis_Framework.xlsx.

This version is robust to NotebookLM outputs saved as Markdown that may contain:
- escaped underscores (e.g., RUN\_ID)
- escaped numbered items (e.g., 1\.)
- HTML non-breaking spaces (&nbsp;)
- an "outer" BEGIN/END wrapper and an "inner" escaped BEGIN/END wrapper

It writes a NEW Excel file and NEW Session_Index.csv so originals are not overwritten.
"""

import argparse
import csv
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook

# -----------------------------
# Config
# -----------------------------
BEGIN = "---BEGIN_NOTEBOOKLM_RESPONSE---"
END = "---END_NOTEBOOKLM_RESPONSE---"

MAX_ITEMS_PER_BUCKET = 15  # enforce "15 rows at most" rule


# -----------------------------
# Text normalization
# -----------------------------
def normalize_md(text: str) -> str:
    """Normalize common Markdown/HTML artifacts."""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("&nbsp;", " ")
    text = text.replace("\u00a0", " ")       # NBSP char
    text = text.replace("\\_", "_")          # escaped underscore
    text = text.replace("\\.", ".")          # escaped dot
    return text


def extract_response_block(full_text: str) -> str:
    """Extract text between BEGIN/END markers. If missing, return full text."""
    t = normalize_md(full_text)
    if BEGIN in t and END in t:
        block = t.split(BEGIN, 1)[1].split(END, 1)[0].strip()
    else:
        block = t.strip()

    # Trim anything before the first numbered item (in case header metadata got pasted inside)
    m = re.search(r"(?m)^\s*(?:\*\*\s*)?(\d+)\.\s+", block)
    if m:
        block = block[m.start():].strip()

    return block


def split_numbered_items(block: str) -> List[Tuple[int, str]]:
    """Split into numbered items. Supports bolded numbering like **1. Title**."""
    block = normalize_md(block).strip()

    pattern = re.compile(r"(?m)^\s*(?:\*\*\s*)?(\d+)\.\s+")
    matches = list(pattern.finditer(block))
    if not matches:
        raise ValueError("No numbered items found. Expected '1.', '2.', etc.")

    items: List[Tuple[int, str]] = []
    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(block)
        num = int(m.group(1))
        item_text = block[m.end():end].strip()
        items.append((num, item_text))

    return items


def strip_md(s: str) -> str:
    return s.replace("**", "").replace("*", "").strip()


def cleanup_lines(text: str) -> str:
    """Remove leading bullets for easier field parsing."""
    t = normalize_md(text)
    out = []
    for line in t.split("\n"):
        line = line.strip()
        line = re.sub(r"^[\u2022\-\*\u25e6]+\s+", "", line)  # • - * ◦
        out.append(line)
    return "\n".join(out).strip()


def find_field(item_text: str, label: str) -> str:
    """
    Extract label-based fields like:
      **Description:** ...
      Description: ...
    """
    t = cleanup_lines(item_text)

    # allow underscore or space variants in label
    label_pattern = re.escape(label).replace("_", "[_ ]")

    # capture until next "SomeLabel:" style line
    pat = re.compile(
        rf"(?is)(?:^|\n)\s*(?:\*\*\s*)?{label_pattern}(?:\s*\*\*)?\s*:\s*(.*?)(?=\n\s*(?:\*\*\s*)?[A-Za-z0-9][A-Za-z0-9_ /()\-]*\s*(?:\*\*)?\s*:\s|\Z)"
    )
    m = pat.search(t)
    return strip_md(m.group(1)) if m else ""


def find_quotes_anywhere(text: str) -> List[str]:
    """Extract all quoted snippets '\"...\"' from a block."""
    t = normalize_md(text)
    return [q.strip() for q in re.findall(r'"([^"]+)"', t)]


def today_str() -> str:
    return datetime.now().strftime("%Y-%m-%d")


# -----------------------------
# Excel helpers
# -----------------------------
def get_headers(ws, header_row=11) -> List[str]:
    headers = []
    col = 1
    while True:
        v = ws.cell(header_row, col).value
        if v is None:
            break
        headers.append(str(v).strip())
        col += 1
    return headers


def find_next_empty_row(ws, header_row=11) -> int:
    row = header_row + 1
    while ws.cell(row, 1).value not in (None, ""):
        row += 1
    return row


def append_row(ws, headers: List[str], row_dict: Dict[str, str], header_row=11) -> int:
    r = find_next_empty_row(ws, header_row=header_row)
    for c, h in enumerate(headers, start=1):
        ws.cell(r, c).value = row_dict.get(h, "")
    return r


# -----------------------------
# Bucket builders (map NotebookLM → Excel columns)
# -----------------------------
def build_cross_cutting_theme_row(run_id: str, pass_id: str, num: int, item_text: str) -> Dict[str, str]:
    title = strip_md(item_text.split("\n", 1)[0])
    definition = find_field(item_text, "Definition")
    contexts = find_field(item_text, "Contextual Differences")
    sessions_raw = find_field(item_text, "Session_IDs")
    sessions = [s.strip() for s in sessions_raw.split(",") if s.strip()] if sessions_raw else []

    quotes = find_quotes_anywhere(item_text)
    q1 = quotes[0] if len(quotes) > 0 else ""
    q2 = quotes[1] if len(quotes) > 1 else ""

    conf = "High" if len(sessions) >= 2 else "Medium"

    return {
        "Theme_ID": f"CCT_{num:02d}",
        "Theme_Name": title,
        "Theme_Definition": definition,
        "How_Theme_Shows_Up_Across_Contexts": contexts,
        "Sessions_Appears_In (IDs)": ", ".join(sessions),
        "Confidence (High/Med/Low)": conf,
        "Evidence_1_Session_ID": sessions[0] if sessions else "",
        "Evidence_1_Speaker_Name": "",
        "Evidence_1_Speaker_Role": "",
        "Evidence_1_Quote_or_Excerpt": q1,
        "Evidence_1_Why_This_Evidence_Matters": "",
        "Evidence_2_Session_ID": sessions[1] if len(sessions) > 1 else (sessions[0] if sessions else ""),
        "Evidence_2_Speaker_Name": "",
        "Evidence_2_Speaker_Role": "",
        "Evidence_2_Quote_or_Excerpt": q2,
        "Evidence_2_Why_This_Evidence_Matters": "",
        "Analyst_Notes": f"Imported from INBOX ({run_id}, {pass_id})",
        "Last_Updated": today_str(),
    }


def build_key_learning_row(run_id: str, pass_id: str, num: int, item_text: str) -> Dict[str, str]:
    title = strip_md(item_text.split("\n", 1)[0])
    statement = find_field(item_text, "Concise learning statement")
    sessions_raw = find_field(item_text, "Session_IDs")
    sessions = [s.strip() for s in sessions_raw.split(",") if s.strip()] if sessions_raw else []
    roles = find_field(item_text, "Speaker role(s)")

    quotes = find_quotes_anywhere(item_text)
    q1 = quotes[0] if len(quotes) > 0 else ""
    q2 = quotes[1] if len(quotes) > 1 else ""

    conf = "High" if len(sessions) >= 2 else "Medium"

    return {
        "Key_Learning_ID": f"KL_{num:02d}",
        "Key_Learning_Statement": statement or title,
        "Description_Notes": title if statement else "",
        "Sessions_Appears_In (IDs)": ", ".join(sessions),
        "Confidence (High/Med/Low)": conf,
        "Evidence_1_Session_ID": sessions[0] if sessions else "",
        "Evidence_1_Speaker_Name": "",
        "Evidence_1_Speaker_Role": roles,
        "Evidence_1_Quote_or_Excerpt": q1,
        "Evidence_1_Why_This_Evidence_Matters": "",
        "Evidence_2_Session_ID": sessions[1] if len(sessions) > 1 else (sessions[0] if sessions else ""),
        "Evidence_2_Speaker_Name": "",
        "Evidence_2_Speaker_Role": roles,
        "Evidence_2_Quote_or_Excerpt": q2,
        "Evidence_2_Why_This_Evidence_Matters": "",
        "Analyst_Notes": f"Imported from INBOX ({run_id}, {pass_id})",
        "Last_Updated": today_str(),
    }


def build_recurring_questions_row(run_id: str, pass_id: str, num: int, item_text: str) -> Dict[str, str]:
    title = strip_md(item_text.split("\n", 1)[0])
    cluster_label = find_field(item_text, "Cluster Label")
    underlying = find_field(item_text, "Description of the underlying need or concern")
    sessions_raw = find_field(item_text, "Session_IDs")
    sessions = [s.strip() for s in sessions_raw.split(",") if s.strip()] if sessions_raw else []

    ex_block = find_field(item_text, "Example question excerpts")
    excerpts = find_quotes_anywhere(ex_block) if ex_block else []

    q1 = excerpts[0] if len(excerpts) > 0 else ""
    q2 = excerpts[1] if len(excerpts) > 1 else ""

    conf = "High" if len(sessions) >= 2 else "Medium"

    return {
        "Question_Cluster_ID": f"QAQ_{num:02d}",
        "Cluster_Label": cluster_label or title,
        "Underlying_Need_or_Concern": underlying,
        "What_Audience_Is_Trying_To_Achieve": "",
        "Sessions_Appears_In (IDs)": ", ".join(sessions),
        "Confidence (High/Med/Low)": conf,
        "Evidence_1_Session_ID": sessions[0] if sessions else "",
        "Evidence_1_Question_As_Worded_or_Excerpt": q1,
        "Evidence_1_Who_Asked (if known)": "",
        "Evidence_1_Why_It_Supports_This_Cluster": "",
        "Evidence_2_Session_ID": sessions[1] if len(sessions) > 1 else (sessions[0] if sessions else ""),
        "Evidence_2_Question_As_Worded_or_Excerpt": q2,
        "Evidence_2_Who_Asked (if known)": "",
        "Evidence_2_Why_It_Supports_This_Cluster": "",
        "Analyst_Notes": f"Imported from INBOX ({run_id}, {pass_id})",
        "Last_Updated": today_str(),
    }


def build_unmet_needs_row(run_id: str, pass_id: str, num: int, item_text: str) -> Dict[str, str]:
    unmet_need = strip_md(item_text.split("\n", 1)[0])
    desc = find_field(item_text, "Description of the unmet need")
    who = find_field(item_text, "Who appears to have the need")
    sessions_raw = find_field(item_text, "Session_IDs")
    sessions = [s.strip() for s in sessions_raw.split(",") if s.strip()] if sessions_raw else []
    evidence = find_field(item_text, "Evidence from transcripts")
    why_matters = find_field(item_text, "Why this unmet need matters for future programming")

    quotes = find_quotes_anywhere(item_text)
    q1 = quotes[0] if len(quotes) > 0 else ""
    q2 = quotes[1] if len(quotes) > 1 else ""

    conf = "High" if len(sessions) >= 2 else "Medium"

    return {
        "Unmet_Need_ID": f"UN_{num:02d}",
        "Unmet_Need": unmet_need,
        "Unmet_Need_Description (Optional detail/clarifier)": desc,
        "Who_Has_This_Need (Stakeholder group)": who,
        "Sessions_Appears_In (IDs)": ", ".join(sessions),
        "Confidence (High/Med/Low)": conf,
        "Evidence_1_Session_ID": sessions[0] if sessions else "",
        "Evidence_1_Speaker_Name (if known)": "",
        "Evidence_1_Speaker_Role": "",
        "Evidence_1_Quote_or_Excerpt": q1 or evidence,
        "Evidence_1_Why_It_Supports_This_Unmet_Need": "",
        "Evidence_2_Session_ID": sessions[1] if len(sessions) > 1 else (sessions[0] if sessions else ""),
        "Evidence_2_Speaker_Name (if known)": "",
        "Evidence_2_Speaker_Role": "",
        "Evidence_2_Quote_or_Excerpt": q2,
        "Evidence_2_Why_It_Supports_This_Unmet_Need": "",
        "Analyst_Notes": why_matters or f"Imported from INBOX ({run_id}, {pass_id})",
        "Last_Updated": today_str(),
    }


def build_missing_elements_row(run_id: str, pass_id: str, num: int, item_text: str) -> Dict[str, str]:
    missing_element = strip_md(item_text.split("\n", 1)[0])
    desc = find_field(item_text, "Description")
    cond = find_field(item_text, "Condition(s) Met")
    evidence = find_field(item_text, "Evidence")
    why = find_field(item_text, "Why this matters")

    quotes = find_quotes_anywhere(item_text)
    q1 = quotes[0] if len(quotes) > 0 else ""
    q2 = quotes[1] if len(quotes) > 1 else ""

    # Topic area heuristic (optional; safe default = blank)
    topic = ""
    low = (missing_element + " " + desc).lower()
    if any(k in low for k in ["policy", "governance", "data access"]):
        topic = "Governance"
    elif any(k in low for k in ["training", "workshop"]):
        topic = "Training"
    elif any(k in low for k in ["use case", "assignment", "rubric"]):
        topic = "Use Cases"
    elif any(k in low for k in ["workforce", "employer", "partnership"]):
        topic = "Workforce"
    elif any(k in low for k in ["evaluate", "evaluation", "impact", "metrics"]):
        topic = "Evaluation"

    return {
        "Missing_Element_ID": f"ME_{num:02d}",
        "Missing_Element": missing_element,
        "Missing_Element_Description (Optional detail/clarifier)": desc,
        "Condition_Met (1-4)": cond,
        "Topic_Area (Governance/Training/Use Cases/Workforce/Evaluation/Other)": topic or "Other",
        "Sessions_Implicated (IDs)": "",  # not always explicit; can be derived later
        "Confidence (High/Med/Low)": "Medium",
        "Evidence_1_Session_ID": "",
        "Evidence_1_Who (Speaker/Role)": "",
        "Evidence_1_Quote_or_Excerpt": q1 or evidence,
        "Evidence_1_Why_It_Shows_This_Is_Missing": "",
        "Evidence_2_Session_ID": "",
        "Evidence_2_Who (Speaker/Role)": "",
        "Evidence_2_Quote_or_Excerpt": q2,
        "Evidence_2_Why_It_Shows_This_Is_Missing": "",
        "Why_This_Matters_For_Future_Programming": why,
        "Suggested_Future_Session_or_Resource": "",
        "Analyst_Notes": f"Imported from INBOX ({run_id}, {pass_id})",
        "Last_Updated": today_str(),
    }


def build_use_cases_row(run_id: str, pass_id: str, num: int, item_text: str) -> Dict[str, str]:
    use_case = strip_md(item_text.split("\n", 1)[0])
    who = find_field(item_text, "Who is using AI")
    context = find_field(item_text, "Context and purpose")
    benefits = find_field(item_text, "Benefits or outcomes described")
    risks = find_field(item_text, "Risks or constraints mentioned")
    session_role = find_field(item_text, "Session_ID and speaker role")
    supporting = find_field(item_text, "Supporting quotes")
    exemplar = find_field(item_text, "Exemplar Status")

    # Exemplar parsing
    is_exemplar = ""
    why_exemplar = ""
    if exemplar:
        # often starts with "Yes." / "No." / "Partial."
        m = re.match(r"(?is)\s*(yes|no|partial)\.?\s*(.*)$", strip_md(exemplar))
        if m:
            is_exemplar = "Y" if m.group(1).lower() == "yes" else "N"
            if m.group(1).lower() == "partial":
                is_exemplar = "N"  # Excel expects Y/N; keep detail in Why_Exemplar
            why_exemplar = m.group(2).strip()
        else:
            # fallback
            is_exemplar = "Y" if "yes" in exemplar.lower() else "N"
            why_exemplar = exemplar

    # Pull session id like "FS02" if present
    sid = ""
    m2 = re.search(r"(FS\d+)", session_role)
    if m2:
        sid = m2.group(1)

    quotes = find_quotes_anywhere(item_text)
    q1 = quotes[0] if len(quotes) > 0 else (supporting or "")
    q2 = quotes[1] if len(quotes) > 1 else ""

    return {
        "Use_Case_ID": f"UC_{num:02d}",
        "Use_Case": use_case,
        "Is_Exemplar (Y/N)": is_exemplar,
        "Why_Exemplar (if Y)": why_exemplar,
        "Intended_User (Role)": who,
        "Context": context,
        "Purpose / Job-to-be-done": context,
        "Workflow_Steps (as described)": "",
        "Tools_or_Systems_Mentioned": "",
        "Benefits_or_Outcomes_Claimed": benefits,
        "Risks_or_Constraints_Mentioned": risks,
        "Sessions_Appears_In (IDs)": sid,
        "Confidence (High/Med/Low)": "Medium",
        "Evidence_1_Session_ID": sid,
        "Evidence_1_Who (Speaker/Role)": session_role,
        "Evidence_1_Quote_or_Excerpt": q1,
        "Evidence_1_Why_It_Supports_This_Use_Case": "",
        "Evidence_2_Session_ID": sid,
        "Evidence_2_Who (Speaker/Role)": session_role,
        "Evidence_2_Quote_or_Excerpt": q2,
        "Evidence_2_Why_It_Supports_This_Use_Case": "",
        "Clip_Candidate (Describe moment; no timestamps yet)": "",
        "Suggested_Future_Session_Topic": "",
        "Analyst_Notes": f"Imported from INBOX ({run_id}, {pass_id})",
        "Last_Updated": today_str(),
    }


# -----------------------------
# Main pipeline
# -----------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--run_id", required=True)
    ap.add_argument("--pass_id", default="PASS_1")
    ap.add_argument("--project_root", default=".")
    ap.add_argument("--inbox_dir", default=None, help="Optional override INBOX folder")
    args = ap.parse_args()

    project_root = Path(args.project_root).resolve()
    run_id = args.run_id
    pass_id = args.pass_id

    excel_path = project_root / "Qual_Analysis_Framework.xlsx"
    session_index_path = project_root / "Session_Index.csv"

    if args.inbox_dir:
        inbox_dir = Path(args.inbox_dir).resolve()
    else:
        inbox_dir = project_root / "03_NotebookLM_Exports" / "INBOX"
        if not inbox_dir.exists():
            # fallback: current directory
            inbox_dir = project_root

    if not excel_path.exists():
        raise FileNotFoundError(f"Missing: {excel_path}")
    if not session_index_path.exists():
        raise FileNotFoundError(f"Missing: {session_index_path}")
    if not inbox_dir.exists():
        raise FileNotFoundError(f"Missing INBOX dir: {inbox_dir}")

    bucket_parsers = [
        ("Cross_Cutting_Themes", build_cross_cutting_theme_row),
        ("Key_Learnings", build_key_learning_row),
        ("Recurring_Audience_Questions", build_recurring_questions_row),
        ("Unmet_Needs", build_unmet_needs_row),
        ("Missing_Elements", build_missing_elements_row),
        ("Use_Cases_Exemplars", build_use_cases_row),
    ]

    wb = load_workbook(excel_path)
    processed = []

    for bucket_name, builder in bucket_parsers:
        inbox_file = inbox_dir / f"{run_id}__{pass_id}__{bucket_name}__NotebookLM_Response.md"
        if not inbox_file.exists():
            raise FileNotFoundError(f"Missing response file: {inbox_file.name}")

        raw = inbox_file.read_text(encoding="utf-8")
        block = extract_response_block(raw)
        items = split_numbered_items(block)

        # enforce max items
        if len(items) > MAX_ITEMS_PER_BUCKET:
            items = items[:MAX_ITEMS_PER_BUCKET]

        if bucket_name not in wb.sheetnames:
            raise KeyError(f"Excel missing expected sheet: {bucket_name}")

        ws = wb[bucket_name]
        headers = get_headers(ws, header_row=11)

        for num, item_text in items:
            row = builder(run_id, pass_id, num, item_text)
            append_row(ws, headers, row, header_row=11)

        processed.append(f"{bucket_name} ({len(items)})")
        print(f"[OK] {bucket_name}: appended {len(items)} row(s)")

    out_excel = project_root / f"Qual_Analysis_Framework__UPDATED_{run_id}.xlsx"
    wb.save(out_excel)
    print(f"[OK] Saved: {out_excel}")

    # Update Session_Index.csv
    out_csv = project_root / f"Session_Index__UPDATED_{run_id}.csv"
    with session_index_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = reader.fieldnames or []

    if "NotebookLM_Status" not in fieldnames:
        fieldnames.append("NotebookLM_Status")

    status_value = f"Pass1 registered to Excel ({run_id})"
    for r in rows:
        r["NotebookLM_Status"] = status_value

    with out_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"[OK] Saved: {out_csv}")
    print(f"[DONE] Buckets processed: {', '.join(processed)}")


if __name__ == "__main__":
    main()
